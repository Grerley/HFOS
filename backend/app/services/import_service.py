"""Excel workbook import pipeline (requirements §16, backlog HFOS-140..142).

Responsibilities:
  * Classify each sheet (monthly budget / scenario / receivables / bonus / other).
  * Parse monthly sheets into owners, sections and line items.
  * Map owner columns → household members and section headers → product categories.
  * Import idempotently (re-running does not duplicate historical periods).
  * Produce a reconciliation report and a review queue for unmapped rows.

Parsing is deliberately tolerant: the source is a human-maintained spreadsheet.
"""
from __future__ import annotations

import io
import re
from calendar import monthrange
from datetime import date

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.enums import CategoryType
from app.models import BudgetLine, BudgetPeriod, Category, HouseholdMember
from app.services import audit_service

# Canonical workbook section headers → default taxonomy section name + type.
SECTION_MAP: dict[str, tuple[str, str]] = {
    "total income": ("Income", CategoryType.INCOME.value),
    "mandatory obligations": ("Mandatory Obligations", CategoryType.EXPENSE.value),
    "insurance": ("Insurance", CategoryType.EXPENSE.value),
    "living expenses": ("Living Expenses", CategoryType.EXPENSE.value),
    "property shortfalls": ("Property Shortfalls", CategoryType.EXPENSE.value),
    "savings & investments": ("Savings & Investments", CategoryType.SAVING.value),
    "ad-hoc expenses": ("Ad hoc Expenses", CategoryType.EXPENSE.value),
    "ad hoc expenses": ("Ad hoc Expenses", CategoryType.EXPENSE.value),
    "discretionary": ("Discretionary", CategoryType.EXPENSE.value),
}

SUMMARY_ROWS = {"total income", "total expenses", "nett", "net"}

MONTHS = {
    m: i
    for i, m in enumerate(
        ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"],
        start=1,
    )
}


def _to_cents(value) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return round(float(value) * 100)
    return None


def classify_sheet(name: str, ws) -> tuple[str, float]:
    """Return (kind, confidence). Kinds: monthly, scenario, receivables, bonus, other."""
    n = name.lower()
    if "scenario" in n or "retirement" in n or "frontier" in n or "bryanston" in n:
        return "scenario", 0.9
    if "salary owed" in n or "owed" in n:
        return "receivables", 0.9
    if "bonus" in n:
        return "bonus", 0.9
    if "carsale" in n or "car sale" in n:
        return "asset_sale", 0.8
    if re.match(r"^[a-z]{3}4[a-z]{3}", n):  # Jan4Feb / Dec4Jan pattern
        return "monthly", 0.95
    # Content signal: an income header in column A.
    for row in ws.iter_rows(min_row=1, max_row=6, max_col=1, values_only=True):
        if row and isinstance(row[0], str) and "total income" in row[0].lower():
            return "monthly", 0.7
    return "other", 0.3


def parse_month_dates(label: str) -> tuple[date, date] | None:
    """Best-effort start/end dates from a label like 'Jan4Feb_2025' or 'Dec4Jan26'."""
    m = re.match(r"^([a-z]{3})4([a-z]{3})", label.lower())
    if not m:
        return None
    start_m, end_m = MONTHS.get(m.group(1)), MONTHS.get(m.group(2))
    if not start_m or not end_m:
        return None
    yr = re.search(r"(20\d{2}|\b\d{2}\b)", label)
    year = 2025
    if yr:
        y = int(yr.group(1))
        year = y if y > 100 else 2000 + y
    start = date(year, start_m, 1)
    end_year = year + 1 if end_m < start_m else year
    end = date(end_year, end_m, monthrange(end_year, end_m)[1])
    return start, end


def parse_monthly_sheet(ws) -> dict:
    """Extract owners, sections and line items from a monthly sheet."""
    # Header row: find the row containing owner names (col C onward) + a Total column.
    owners: list[tuple[int, str]] = []  # (column_index, owner_name)
    header_row = 1
    for r in range(1, 4):
        vals = [ws.cell(row=r, column=c).value for c in range(3, 7)]
        if any(isinstance(v, str) and v.strip() for v in vals):
            header_row = r
            for c in range(3, 7):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str) and v.strip() and v.strip().lower() != "total":
                    owners.append((c, v.strip()))
            break

    sections: list[dict] = []
    current: dict | None = None
    reconciliation = {"total_income": None, "total_expenses": None, "nett": None}

    for r in range(header_row + 1, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        if not isinstance(label, str) or not label.strip():
            continue
        key = label.strip().lower()
        total = _to_cents(ws.cell(row=r, column=6).value)

        if key in SUMMARY_ROWS:
            # "Total Income" appears twice in the workbook (summary row + section
            # header); keep the first non-empty value so the header row's empty
            # total does not overwrite the real summary figure.
            if key == "total income" and total is not None and reconciliation["total_income"] is None:
                reconciliation["total_income"] = total
            elif key == "total expenses" and total is not None:
                reconciliation["total_expenses"] = total
            elif key in ("nett", "net") and total is not None:
                reconciliation["nett"] = total
            continue

        matched_section = next((k for k in SECTION_MAP if key.startswith(k)), None)
        if matched_section:
            current = {
                "raw_name": label.strip(),
                "mapped": SECTION_MAP[matched_section],
                "lines": [],
            }
            sections.append(current)
            continue

        # A line row: capture per-owner amounts and the due-date note.
        due_note = ws.cell(row=r, column=2).value
        amounts = {}
        for col, owner_name in owners:
            cents = _to_cents(ws.cell(row=r, column=col).value)
            if cents:
                amounts[owner_name] = cents
        line = {
            "raw_name": label.strip(),
            "due_note": str(due_note) if isinstance(due_note, str) else None,
            "amounts": amounts,
            "total_cents": total or sum(amounts.values()),
            "cell": f"A{r}",
        }
        if current is None:
            # Lines before the first recognised section (income block).
            current = {"raw_name": "Income", "mapped": ("Income", CategoryType.INCOME.value),
                       "lines": []}
            sections.append(current)
        current["lines"].append(line)

    return {"owners": [o[1] for o in owners], "sections": sections,
            "reconciliation": reconciliation}


def analyze_workbook(file_bytes: bytes) -> dict:
    """Preview: classify sheets and summarise monthly content, without importing."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    sheets = []
    owner_names: set[str] = set()
    for name in wb.sheetnames:
        ws = wb[name]
        kind, conf = classify_sheet(name, ws)
        entry = {"sheet": name, "kind": kind, "confidence": conf}
        if kind == "monthly":
            parsed = parse_monthly_sheet(ws)
            owner_names.update(parsed["owners"])
            entry["owners"] = parsed["owners"]
            entry["line_count"] = sum(len(s["lines"]) for s in parsed["sections"])
            entry["dates"] = None
            d = parse_month_dates(name)
            entry["dates"] = {"start": d[0].isoformat(), "end": d[1].isoformat()} if d else None
        sheets.append(entry)
    wb.close()
    return {"sheets": sheets, "detected_owners": sorted(owner_names)}


def import_workbook(
    db: Session,
    *,
    household_id: int,
    file_bytes: bytes,
    owner_mapping: dict[str, int] | None = None,
    actor_user_id: int | None = None,
    limit_sheets: list[str] | None = None,
) -> dict:
    """Idempotently import monthly sheets as budget periods + lines.

    owner_mapping maps workbook owner names → household member ids. Unknown owners
    are auto-created as planning-only members. Returns a reconciliation report.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    owner_mapping = dict(owner_mapping or {})

    categories = {
        (c.name.lower()): c
        for c in db.scalars(select(Category).where(Category.household_id == household_id)).all()
    }
    members = {
        m.name.lower(): m
        for m in db.scalars(
            select(HouseholdMember).where(HouseholdMember.household_id == household_id)
        ).all()
    }

    report = {
        "periods_imported": 0,
        "periods_skipped": 0,
        "lines_imported": 0,
        "rows_skipped_empty": 0,
        "review_queue": [],
        "reconciliation": [],
        "sheets": [],
    }

    def resolve_member(owner_name: str) -> int:
        if owner_name in owner_mapping:
            return owner_mapping[owner_name]
        m = members.get(owner_name.lower())
        if m:
            return m.id
        created = HouseholdMember(
            household_id=household_id, name=owner_name, relationship_label="imported",
            role="partner",
        )
        db.add(created)
        db.flush()
        members[owner_name.lower()] = created
        owner_mapping[owner_name] = created.id
        return created.id

    def resolve_category(section_name: str, ctype: str) -> Category:
        c = categories.get(section_name.lower())
        if c:
            return c
        c = Category(household_id=household_id, name=section_name, type=ctype, is_section=True)
        db.add(c)
        db.flush()
        categories[section_name.lower()] = c
        return c

    for name in wb.sheetnames:
        if limit_sheets and name not in limit_sheets:
            continue
        ws = wb[name]
        kind, _ = classify_sheet(name, ws)
        if kind != "monthly":
            report["sheets"].append({"sheet": name, "kind": kind, "status": "not_imported"})
            continue

        dates = parse_month_dates(name)
        if not dates:
            report["review_queue"].append(
                {"sheet": name, "reason": "could_not_map_dates", "action": "confirm period dates"}
            )
            report["sheets"].append({"sheet": name, "kind": kind, "status": "needs_date_mapping"})
            continue

        # Idempotency: skip if a period with this label already exists.
        existing = db.scalar(
            select(BudgetPeriod).where(
                BudgetPeriod.household_id == household_id, BudgetPeriod.label == name
            )
        )
        if existing:
            report["periods_skipped"] += 1
            report["sheets"].append({"sheet": name, "kind": kind, "status": "already_imported"})
            continue

        parsed = parse_monthly_sheet(ws)
        period = BudgetPeriod(
            household_id=household_id, label=name, start_date=dates[0], end_date=dates[1],
            status="closed", source="workbook_import",
        )
        db.add(period)
        db.flush()

        line_count = 0
        for section in parsed["sections"]:
            mapped_name, ctype = section["mapped"]
            category = resolve_category(mapped_name, ctype)
            for line in section["lines"]:
                if not line["amounts"] and not line["total_cents"]:
                    report["rows_skipped_empty"] += 1
                    # Zero lines are kept (planned future categories) but flagged compactly.
                owner_id = None
                if line["amounts"]:
                    owner_name = max(line["amounts"], key=line["amounts"].get)
                    owner_id = resolve_member(owner_name)
                bl = BudgetLine(
                    period_id=period.id,
                    household_id=household_id,
                    category_id=category.id,
                    item_name=line["raw_name"],
                    owner_member_id=owner_id,
                    planned_amount_cents=line["total_cents"] or 0,
                    actual_amount_cents=0,
                    due_note=line["due_note"],
                    recurrence="monthly",
                    payment_status="planned",
                    source_ref=f"workbook:{name}!{line['cell']}",
                    needs_review=(line["total_cents"] or 0) == 0,
                )
                db.add(bl)
                line_count += 1

        db.flush()
        # Reconcile imported income/expense totals against the sheet's own totals.
        from app.services.queries import load_lines_for_calc
        from app.services import calculations as calc

        calc_lines = load_lines_for_calc(db, household_id, period.id)
        rec = parsed["reconciliation"]
        report["reconciliation"].append(
            {
                "sheet": name,
                "workbook_total_income_cents": rec["total_income"],
                "imported_total_income_cents": calc.total_income(calc_lines),
                "workbook_total_expenses_cents": rec["total_expenses"],
                "imported_total_expenses_cents": calc.total_expenses(calc_lines),
            }
        )
        report["periods_imported"] += 1
        report["lines_imported"] += line_count
        report["sheets"].append(
            {"sheet": name, "kind": kind, "status": "imported", "lines": line_count}
        )

    wb.close()
    audit_service.record(
        db,
        action="workbook.import",
        entity_type="household",
        entity_id=household_id,
        household_id=household_id,
        actor_user_id=actor_user_id,
        detail={k: report[k] for k in ("periods_imported", "periods_skipped", "lines_imported")},
    )
    return report
