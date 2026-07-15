"""Generate a synthetic sample workbook that mirrors the real family workbook's
structure (owner columns, sectioned categories, due-date notes, ZAR) WITHOUT any
real personal data. Used for the import demo and tests.

Run: python -m app.scripts.gen_sample_workbook [output_path]
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

OWNERS = ["Alex", "Sam", "Robin"]

# section -> list of (item, due_note, {owner: rand_amount})
MONTHLY_TEMPLATE = {
    "__income__": [
        ("Monthly salary", "25th of every month", {"Alex": 52000}),
        ("Business income", "1st of every month", {"Sam": 31000, "Robin": 180000}),
        ("Rental income", "1st of every month", {"Robin": 12000}),
    ],
    "Mandatory Obligations": [
        ("Bond", "3rd of every month", {"Robin": 28500}),
        ("Utilities (water, electricity, gas)", None, {"Alex": 4200}),
        ("Vehicle / transport", None, {"Robin": 9800}),
        ("Security", None, {"Alex": 620}),
        ("Internet", None, {"Alex": 899, "Robin": 1100}),
        ("Phone", None, {"Alex": 1100, "Robin": 5200}),
        ("Domestic worker", None, {"Sam": 3600}),
        ("Bank charges", None, {"Alex": 410}),
        ("School fees", "1st of every month", {"Robin": 8500}),
    ],
    "Insurance": [
        ("Car insurance", None, {"Alex": 2450}),
        ("Home contents", None, {"Alex": 1300}),
        ("Medical aid", None, {"Alex": 6100}),
        ("Life insurance", None, {"Alex": 2100, "Robin": 3200}),
        ("Funeral policy", None, {"Alex": 640}),
    ],
    "Living Expenses": [
        ("Groceries", None, {"Sam": 11000}),
        ("Subscriptions", None, {"Alex": 720}),
        ("Fuel", None, {"Sam": 3200, "Robin": 3600}),
        ("Entertainment", None, {"Alex": 1400}),
        ("Family support", None, {"Robin": 4000}),
    ],
    "Property Shortfalls": [
        ("Rental unit A shortfall", "8th of every month", {"Robin": 3500}),
        ("Rental unit A utilities", None, {"Robin": 2100}),
    ],
    "Savings & Investments": [
        ("Retirement", None, {"Alex": 7600, "Robin": 8000}),
        ("Emergency fund", None, {"Alex": 2500}),
        ("College fund", None, {"Alex": 2400}),
        ("Investment platform", None, {"Robin": 7500}),
    ],
    "Discretionary": [
        ("Tithe", None, {"Robin": 5000}),
        ("Allowances", None, {"Alex": 1200}),
    ],
}

# Small per-month multipliers so periods differ realistically.
MONTHS = [("Jan4Feb_2025", 1.00), ("Feb4Mar_2025", 1.03), ("Mar4Apr_2025", 0.98)]


def _build_sheet(wb, label: str, factor: float) -> None:
    ws = wb.create_sheet(title=label)
    ws["B1"] = "When"
    for i, owner in enumerate(OWNERS):
        ws.cell(row=1, column=3 + i, value=owner)
    ws.cell(row=1, column=3 + len(OWNERS), value="Total")

    # Summary rows are written last once totals are known; reserve rows 2-4.
    ws["A2"], ws["A3"], ws["A4"] = "Total Income", "Total expenses", "Nett"

    total_income = 0
    total_expenses = 0
    r = 6

    def write_line(name, due, amounts, is_income):
        nonlocal r, total_income, total_expenses
        ws.cell(row=r, column=1, value=name)
        if due:
            ws.cell(row=r, column=2, value=due)
        row_total = 0
        for i, owner in enumerate(OWNERS):
            amt = amounts.get(owner)
            if amt:
                val = round(amt * factor)
                ws.cell(row=r, column=3 + i, value=val)
                row_total += val
        ws.cell(row=r, column=3 + len(OWNERS), value=row_total)
        if is_income:
            total_income += row_total
        else:
            total_expenses += row_total
        r += 1

    # Income block
    ws.cell(row=r, column=1, value="Total Income")
    r += 1
    for name, due, amounts in MONTHLY_TEMPLATE["__income__"]:
        write_line(name, due, amounts, True)

    # Expense/saving sections
    for section, lines in MONTHLY_TEMPLATE.items():
        if section == "__income__":
            continue
        ws.cell(row=r, column=1, value=section)
        r += 1
        for name, due, amounts in lines:
            write_line(name, due, amounts, False)

    ws.cell(row=2, column=6, value=total_income)
    ws.cell(row=3, column=6, value=total_expenses)
    ws.cell(row=4, column=6, value=total_income - total_expenses)


def build_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop default sheet
    for label, factor in MONTHS:
        _build_sheet(wb, label, factor)
    # A scenario sheet so the classifier has something non-monthly to categorise.
    scen = wb.create_sheet(title="Retirement scenario")
    scen["A1"] = "Retirement scenario assumptions"
    scen["A2"], scen["B2"] = "Monthly contribution", 12000
    scen["A3"], scen["B3"] = "Years to retirement", 25
    return wb


def main() -> None:
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else (
        Path(__file__).resolve().parents[2] / "data" / "sample_workbook.xlsx"
    )
    out.parent.mkdir(parents=True, exist_ok=True)
    build_workbook().save(out)
    print(f"Wrote sample workbook → {out}")


if __name__ == "__main__":
    main()
